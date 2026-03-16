from typing import List
from fastapi import HTTPException
import io
from openpyxl import Workbook
from openpyxl.styles import Font

from backend.app.auth.models import User
from backend.app.core.logging import get_logger

logger = get_logger()


def create_list_users_excel(list_user: List[User]) -> io.BytesIO:
    """Tạo file Excel danh sách người dùng"""
    try:
        wb = Workbook()
        ws = wb.active
        if ws:
            ws.title = "List Users"
        else:
            ws = wb.create_sheet("List Users")

        headers = [
            "ID",
            "Username",
            "Email",
            "First Name",
            "Last Name",
            "Created At",
            "Updated At",
        ]

        if ws:
            ws.append(headers)

        if ws:
            for cell in ws[1]:
                if cell:
                    cell.font = Font(bold=True)

        for user in list_user:
            if ws:
                ws.append([
                str(user.id),
                user.username,
                user.email,
                user.first_name,
                user.last_name,
                user.created_at.strftime("%Y-%m-%d %H:%M:%S")
                if user.created_at else "",
                user.updated_at.strftime("%Y-%m-%d %H:%M:%S")
                if user.updated_at else "",
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    except Exception as e:
        logger.error(f"Lỗi khi tạo file Excel danh sách người dùng: {e}")
        raise HTTPException(
            status_code=500,
            detail="Lỗi khi tạo file Excel danh sách người dùng"
        )
